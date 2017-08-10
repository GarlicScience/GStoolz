from openpyxl import Workbook, load_workbook
from core.data.game.net.net_pickle import NetPickle
import pandas as pd

class PickleToXL:
    wb = None
    @classmethod
    def make_excel(cls):
        c_frame, b_frame, mes = NetPickle.load_as_frames()
        wb = load_workbook('template.xlsx')
        cls.wb = wb

        prices_sheet = wb.active
        prices_sheet.title = 'Prices + bundles'

        cls.fill_levels_sheet(wb['Level'], c_frame)

        base_contracts_sheet = wb.create_sheet('Base contracts')
        cls.fill_base_contracts_sheet(base_contracts_sheet, c_frame)

        contracts_sheet = wb.create_sheet('Contracts')
        cls.fill_contracts_sheet(contracts_sheet, c_frame)

        shuttle_contracts_sheet = wb.create_sheet('Shuttle contracts')
        cls.fill_shuttle_contracts_sheet(shuttle_contracts_sheet, c_frame)

        cls.fill_contracts_sheet(contracts_sheet, c_frame)

        wb.save('balance.xlsx')

    @classmethod
    def fill_levels_sheet(cls, levels_sheet, c_frame):
        levels_sheet['C1'] = 'Water_max'
        levels_sheet['D1'] = 'Air_max'
        levels_sheet['E1'] = 'Energy_max'

        c_frame = c_frame[(c_frame['name'] == 'Water') |
                          (c_frame['name'] == 'Air') |
                          (c_frame['name'] == 'Energy')].reset_index()

        del c_frame['index']

        water_max = eval(c_frame['max_amount'][0])
        air_max = eval(c_frame['max_amount'][1])
        energy_max = eval(c_frame['max_amount'][2])

        for wat in water_max:
            levels_sheet['C{}'.format(int(wat['plvl']) + 1)] = wat['amount']

        for wat in air_max:
            levels_sheet['D{}'.format(int(wat['plvl']) + 1)] = wat['amount']

        for wat in energy_max:
            levels_sheet['E{}'.format(int(wat['plvl']) + 1)] = wat['amount']


    @classmethod
    def __r(cls, a):
        return "ROUND({}, 5)".format(a)

    @classmethod
    def fill_base_contracts_sheet(cls, base_contracts_sheet, c_frame):

        base_contracts_sheet['A1'] = 'id'
        base_contracts_sheet['B1'] = 'name'
        base_contracts_sheet['C1'] = 'recovery_time'
        base_contracts_sheet['D1'] = 'self_cost_$'

        c_frame = c_frame[(c_frame['type'] == 'base') &
                          (c_frame['name'] != 'Currency')].reset_index()

        del c_frame['index']

        for i in range(len(c_frame['id'])):
            base_contracts_sheet['A' + str(i + 2)] = c_frame['id'][i]
            base_contracts_sheet['B' + str(i + 2)] = c_frame['name'][i]
            base_contracts_sheet['C' + str(i + 2)] = c_frame['recovery_time'][i]
            base_contracts_sheet['D' + str(i + 2)] = "={}*{}".format(cls.__r("'Prices + bundles'!B5")
                                                                     ,cls.__r("C" + str(i+2)))

    @classmethod
    def fill_shuttle_contracts_sheet(cls, shuttle_contracts_sheet, c_frame):
        shuttle_contracts_sheet['A1'] = 'id'
        shuttle_contracts_sheet['B1'] = 'name'
        shuttle_contracts_sheet['C1'] = 'sell_cost_for_one'
        shuttle_contracts_sheet['D1'] = 'donate_cost_for_one'
        shuttle_contracts_sheet['E1'] = 'chance'
        shuttle_contracts_sheet['F1'] = 'roll_count'
        shuttle_contracts_sheet['G1'] = 'self_cost_$'
        rew_list = c_frame[c_frame['id'] == 48].reset_index()
        rew_list = eval(rew_list['random_rewards'][0])
        rew_list = list(rew_list.values())[1:]

        c_frame = c_frame[c_frame['type'] == 'space'].reset_index()

        del c_frame['index']

        for i in range(len(c_frame['id'])):
            shuttle_contracts_sheet['A' + str(i + 2)] = c_frame['id'][i]
            shuttle_contracts_sheet['B' + str(i + 2)] = c_frame['name'][i]
            shuttle_contracts_sheet['C' + str(i + 2)] = c_frame['sell_cost_for_one'][i]
            shuttle_contracts_sheet['D' + str(i + 2)] = c_frame['donate_cost_for_one'][i]
            shuttle_contracts_sheet['E' + str(i + 2)] =\
                [reward for reward in rew_list if str(reward['id']) == str(c_frame['id'][i])][0]['chance'].replace('.',',')
            shuttle_contracts_sheet['F' + str(i + 2)] = 1
            cell, a = cls.get_contract('48')
            shuttle_contracts_sheet['G' + str(i + 2)] = "={}*(1/({}*{}))".format(cell,
                                                                           'F' + str(i + 2),
                                                                           cls.__r('E' + str(i + 2)))


    @classmethod
    def fill_contracts_sheet(cls, contracts_sheet, c_frame):

        contracts_sheet['A1'] = 'id'
        contracts_sheet['B1'] = 'name'
        contracts_sheet['C1'] = 'time'
        contracts_sheet['D1'] = 'needs_lvl'
        contracts_sheet['E1'] = 'get_exp'
        contracts_sheet['F1'] = 'sell_cost_for_one'
        contracts_sheet['G1'] = 'donate_cost_for_one'
        contracts_sheet['H1'] = 'needs_contracts'
        contracts_sheet['I1'] = 'self_cost_$'

        c_frame = c_frame[(c_frame['type'] != 'base') & (c_frame['type'] != 'space')].reset_index()

        del c_frame['index']

        for i in range(len(c_frame['id'])):
            contracts_sheet['A' + str(i + 2)] = c_frame['id'][i]
            contracts_sheet['B' + str(i + 2)] = c_frame['name'][i]
            contracts_sheet['C' + str(i + 2)] = c_frame['time'][i]
            contracts_sheet['D' + str(i + 2)] = c_frame['needs_lvl'][i]
            contracts_sheet['E' + str(i + 2)] = c_frame['get_exp'][i]
            contracts_sheet['F' + str(i + 2)] = c_frame['sell_cost_for_one'][i]
            contracts_sheet['G' + str(i + 2)] = c_frame['donate_cost_for_one'][i]

        for i in range(len(c_frame['id'])):
            needs_c = ""
            needs_contracts = ""
            formula = ""
            if len(c_frame['needs_contracts'][i]) > 0:
                needs_c = eval(c_frame['needs_contracts'][i])
                for req in needs_c:
                    cost_cell, name = cls.get_contract(req['id'])
                    needs_contracts += name + ' ({}), '.format(req['amount'])
                    formula += "({}*{})+".format(cls.__r(cost_cell), req['amount'])

            formula = formula[:-1]
            needs_contracts = needs_contracts[:-2]

            contracts_sheet['H' + str(i + 2)] = needs_contracts
            contracts_sheet['I' + str(i + 2)] = "={}*{}+{}".format(cls.__r("'Prices + bundles'!B5"),
                                                                   'C' + str(i + 2),
                                                                   formula)

    @classmethod
    def get_contract(cls, id):
        if int(id) == 4:
            return "'Prices + bundles'!Q2", 'Currency'
        elif int(id) == 0:
            return "'Prices + bundles'!F2", 'Crystal'
        elif int(id) <= 5:
            base_sheet = cls.wb['Base contracts']
            cur_cell = 'A2'
            while len(str(base_sheet[cur_cell].value)) > 0:
                if str(base_sheet[cur_cell].value) == str(id):
                    return "'Base contracts'!{}".format('D' + cur_cell[1:]), base_sheet['B' + cur_cell[1:]].value
                cur_cell = 'A' + str(int(cur_cell[1:]) + 1)
        else:
            contracts_sheet = cls.wb['Contracts']
            cur_cell = 'A2'
            while contracts_sheet[cur_cell].value is not None:
                if str(contracts_sheet[cur_cell].value) == str(id):
                    print('found------------------------------')
                    return "'Contracts'!{}".format('I' + cur_cell[1:]), contracts_sheet['B' + cur_cell[1:]].value
                print('found' + cur_cell)
                cur_cell = 'A' + str(int(cur_cell[1:]) + 1)

        try:
            shuttle_sheet = cls.wb['Shuttle contracts']
            cur_cell = 'A2'
            while len(str(shuttle_sheet[cur_cell].value)) > 0:
                if str(shuttle_sheet[cur_cell].value) == str(id):
                    return "'Shuttle contracts'!{}".format('G' + cur_cell[1:]), shuttle_sheet['B' + cur_cell[1:]].value
                cur_cell = 'A' + str(int(cur_cell[1:]) + 1)
        except Exception:
            pass

        return "", ""
